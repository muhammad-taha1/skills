#!/usr/bin/env python3
"""Regenerate expenses.xlsx, the eval fixture.

The fixture is deliberately shaped to expose the failure modes the skill exists to
prevent: live SUM/SUMIF formulas that a careless `data_only=True` round-trip would
flatten to stale values, a fixed category taxonomy an agent might silently extend,
and two uncategorized rows that force a real decision.
"""
from openpyxl import Workbook

CATEGORIES = ["Software", "Travel", "Meals", "Office", "Misc"]

ROWS = [
    ("2026-01-03", "Figma annual seat",        "Software", 144.00),
    ("2026-01-07", "Flight LHR-SFO",           "Travel",   612.40),
    ("2026-01-09", "Team lunch",               "Meals",     87.25),
    ("2026-01-12", "Standing desk riser",      "Office",   210.00),
    ("2026-01-15", "Datadog monthly",          "Software", 380.00),
    ("2026-01-18", "Airport parking",          "Travel",    44.00),
    # Deliberately uncategorized — the agent must fit these into CATEGORIES above,
    # not invent "Conference" or "Legal" as new ones.
    ("2026-01-21", "Renewal - domain + SSL",   None,        68.99),
    ("2026-01-24", "Notary fee for contract",  None,       125.00),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Jan"
    ws.append(["Date", "Description", "Category", "Amount"])
    for row in ROWS:
        ws.append(list(row))

    total_row = len(ROWS) + 2
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{len(ROWS) + 1})")

    summary = wb.create_sheet("Summary")
    summary.append(["Category", "Total"])
    for i, cat in enumerate(CATEGORIES, start=2):
        summary.cell(row=i, column=1, value=cat)
        summary.cell(row=i, column=2,
                     value=f"=SUMIF(Jan!C:C,A{i},Jan!D:D)")

    out = "expenses.xlsx"
    wb.save(out)
    print(f"wrote {out}: {len(ROWS)} rows, "
          f"{1 + len(CATEGORIES)} formula cells, {len(CATEGORIES)} categories")


if __name__ == "__main__":
    main()
