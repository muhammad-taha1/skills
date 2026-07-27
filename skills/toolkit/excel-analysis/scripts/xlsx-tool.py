#!/usr/bin/env python3
"""Deterministic helpers for safe Excel work. Requires openpyxl.

Usage:
  python xlsx-tool.py survey <workbook.xlsx>          # structure report (read-only)
  python xlsx-tool.py backup <workbook.xlsx> [dir]    # timestamped copy, prints path
  python xlsx-tool.py values <workbook.xlsx> <Sheet> [max_rows]  # cached values as TSV

Run `survey` and `backup` before any edit. `survey` never writes.
"""
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed — run: pip install openpyxl")


def survey(path: Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False,
                                keep_vba=path.suffix.lower() == ".xlsm")
    report = {"file": str(path), "sheets": []}
    for ws in wb.worksheets:
        formulas, merged = [], [str(r) for r in ws.merged_cells.ranges]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1), [])]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})
        report["sheets"].append({
            "name": ws.title,
            "dims": ws.dimensions,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "header_row": header,
            "merged_ranges": merged,
            "formula_count": len(formulas),
            # cap output; full list rarely needed and can be huge
            "formulas": formulas[:60],
            "formulas_truncated": len(formulas) > 60,
        })
    report["defined_names"] = sorted(wb.defined_names.keys()) if hasattr(wb, "defined_names") else []
    print(json.dumps(report, indent=1, default=str))


def backup(path: Path, dest_dir: Path) -> None:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = dest_dir / f"{path.stem}.backup-{stamp}{path.suffix}"
    dest_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, dest)
    print(dest)


def values(path: Path, sheet: str, max_rows: int) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"sheet {sheet!r} not found; sheets: {wb.sheetnames}")
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            print(f"... truncated at {max_rows} rows (of {ws.max_row})")
            break
        print("\t".join("" if v is None else str(v) for v in row))


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    cmd, wb_path = sys.argv[1], Path(sys.argv[2])
    if not wb_path.exists():
        sys.exit(f"not found: {wb_path}")
    if cmd == "survey":
        survey(wb_path)
    elif cmd == "backup":
        backup(wb_path, Path(sys.argv[3]) if len(sys.argv) > 3 else wb_path.parent)
    elif cmd == "values":
        values(wb_path, sys.argv[3], int(sys.argv[4]) if len(sys.argv) > 4 else 200)
    else:
        sys.exit(__doc__)
